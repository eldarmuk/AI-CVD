"""
Data loading pipeline: Extract data from Excel files and load into SQLite database.
"""

import pandas as pd
import sqlite3
from pathlib import Path
import logging
from typing import List, Optional
from openpyxl import load_workbook
from src.components.database import initialize_database, normalize_from_sources, bulk_upsert_seniors

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RAW_DATA_PATH = Path(__file__).parent.parent / "data" / "raw" / "HRP_new"


def _load_single_measurement_file(excel_path: Path) -> pd.DataFrame:
    """Load and normalize one measurement Excel file (multi-sheet)."""
    logger.info(f"Loading measurements from {excel_path}")

    xls = pd.ExcelFile(excel_path)
    sheet_names = xls.sheet_names
    logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")

    all_measurements = []

    for sheet_name in sheet_names:
        if sheet_name.startswith("_"):
            continue

        df = pd.read_excel(excel_path, sheet_name=sheet_name, engine="openpyxl")
        logger.info(f"  {sheet_name}: {len(df)} rows, {df.shape[1]} columns")

        df_normalized = pd.DataFrame({
            "senior_id": df.iloc[:, 0],
            "value": df.iloc[:, 1],
            "sbp": df.iloc[:, 2],
            "dbp": df.iloc[:, 3],
            "date": df.iloc[:, 4],
            "type": df.iloc[:, 5]
        })

        # Type conversions
        df_normalized["senior_id"] = pd.to_numeric(df_normalized["senior_id"], errors="coerce")
        df_normalized["value"] = pd.to_numeric(df_normalized["value"], errors="coerce")
        df_normalized["sbp"] = pd.to_numeric(df_normalized["sbp"], errors="coerce")
        df_normalized["dbp"] = pd.to_numeric(df_normalized["dbp"], errors="coerce")
        df_normalized["type"] = df_normalized["type"].astype(str).str.strip()

        # Drop rows missing senior_id
        df_normalized = df_normalized.dropna(subset=["senior_id"])

        all_measurements.append(df_normalized)

    return pd.concat(all_measurements, ignore_index=True)


def load_measurements_data(excel_path: Path | List[Path]) -> pd.DataFrame:
    """Load measurement data from one or many Excel files and normalize."""
    paths: List[Path] = [excel_path] if isinstance(excel_path, Path) else list(excel_path)
    frames = []
    for path in paths:
        frames.append(_load_single_measurement_file(path))
    df_all = pd.concat(frames, ignore_index=True)
    logger.info(f"Total measurements loaded across files: {len(df_all)}")
    return df_all


def load_seniors_demographics(excel_path: Path) -> pd.DataFrame:
    """Load seniors demographics (gender, birthdate, age)."""
    logger.info(f"Loading seniors demographics from {excel_path}")
    df = pd.read_excel(excel_path, engine="openpyxl")
    df = df.rename(columns={
        "seniorID": "senior_id",
        "gender": "gender",
        "birthDate": "birthdate",
        "age": "age"
    })
    if "senior_id" not in df.columns:
        raise ValueError("seniors demographics file must contain 'seniorID' column")

    df["senior_id"] = pd.to_numeric(df["senior_id"], errors="coerce")
    df["age"] = pd.to_numeric(df.get("age"), errors="coerce")
    if "birthdate" in df.columns:
        df["birthdate"] = pd.to_datetime(df["birthdate"], errors="coerce").dt.date

    df = df.dropna(subset=["senior_id"])
    logger.info(f"Loaded {len(df)} senior demographic rows")
    return df[["senior_id", "gender", "birthdate", "age"]]


def insert_seniors_demographics(df: pd.DataFrame, conn: sqlite3.Connection):
    """Upsert seniors with demographic info."""
    if df.empty:
        logger.info("No senior demographics to insert")
        return

    df_clean = df.copy()
    for col in ["gender", "birthdate", "age"]:
        if col in df_clean.columns:
            df_clean[col] = df_clean[col].where(pd.notnull(df_clean[col]), None)

    records = df_clean.to_dict(orient="records")
    with conn:
        conn.executemany(
            """
            INSERT OR REPLACE INTO seniors (id, gender, birthdate, age)
            VALUES (:senior_id, :gender, :birthdate, :age)
            """,
            records,
        )
    logger.info(f"Upserted {len(df)} seniors with demographics")


def load_medical_info(excel_path: Path) -> pd.DataFrame:
    """Load medical information (diseases and medicines)."""
    logger.info(f"Loading medical info from {excel_path}")
    df = pd.read_excel(excel_path, engine="openpyxl")
    
    # Rename columns to match database schema
    df = df.rename(columns={
        'seniorID': 'senior_id',
        'diseaseNames': 'disease_names',
        'medicineNames': 'medicine_names'
    })
    
    # Handle duplicates: keep last entry for each senior_id
    if df['senior_id'].duplicated().any():
        duplicate_count = df['senior_id'].duplicated().sum()
        logger.warning(f"  Found {duplicate_count} duplicate senior_ids - keeping last entry for each")
        df = df.drop_duplicates(subset=['senior_id'], keep='last')
    
    logger.info(f"  Loaded {len(df)} records")
    return df


def load_alerts(excel_path: Path) -> pd.DataFrame:
    """Load SOS alerts data."""
    logger.info(f"Loading alerts from {excel_path}")
    df = pd.read_excel(excel_path, engine="openpyxl")
    
    # Rename columns to match database schema
    df = df.rename(columns={
        'seniorID': 'senior_id',
        'alertDate': 'alert_date',
        'sosNote': 'sos_note'
    })
    
    logger.info(f"  Loaded {len(df)} records")
    return df


def insert_measurements(df: pd.DataFrame, conn: sqlite3.Connection, batch_size: int = 50000):
    """Insert measurements into database with batch processing."""
    logger.info(f"Inserting {len(df)} measurements...")
    
    # Insert measurements in batches
    for i in range(0, len(df), batch_size):
        batch = df.iloc[i:i+batch_size]
        batch.to_sql("measurements", conn, if_exists="append", index=False)
        if (i // batch_size + 1) % 10 == 0:
            logger.info(f"    Inserted {i + batch_size:,} measurements...")
    
    conn.commit()
    logger.info("✓ Measurements inserted successfully")


def insert_measurements_from_excel(
    excel_paths: List[Path],
    conn: sqlite3.Connection,
    batch_rows: int = 100_000,
    sheet_filters: Optional[List[str]] = None,
    resume: bool = True,
) -> None:
    """Stream measurements from one or more Excel files directly into SQLite.

    Processes each worksheet in read-only streaming mode to avoid loading the
    entire dataset into memory. Commits in batches for resilience and speed.
    """
    total_inserted = 0
    cur = conn.cursor()
    for excel_path in excel_paths:
        logger.info(f"Streaming measurements from {excel_path}")
        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                sheet_name = ws.title
                if sheet_name.startswith("_"):
                    continue
                if sheet_filters and sheet_name not in sheet_filters:
                    continue
                source_key = f"measurements::{excel_path.name}::{sheet_name}"
                # Skip if already done (only when resume is enabled)
                if resume:
                    try:
                        row = cur.execute(
                            "SELECT status FROM ingestion_state WHERE source = ?",
                            (source_key,),
                        ).fetchone()
                        if row and row[0] == "done":
                            logger.info(f"✓ Skipping already processed sheet '{sheet_name}'")
                            continue
                    except sqlite3.OperationalError:
                        # ingestion_state may not exist on very old DBs
                        pass

                # TODO: Validate header structure
                # header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
                _ = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
                # Expect first 6 columns: seniorID, value, sbp, dbp, date, type
                rows_buffer: list[dict] = []
                processed = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    rec = {
                        "senior_id": row[0],
                        "value": row[1],
                        "sbp": row[2],
                        "dbp": row[3],
                        "date": row[4],
                        "type": row[5],
                    }
                    rows_buffer.append(rec)
                    if len(rows_buffer) >= batch_rows:
                        df = pd.DataFrame.from_records(rows_buffer, columns=["senior_id","value","sbp","dbp","date","type"])
                        # Type conversions and cleanup
                        df["senior_id"] = pd.to_numeric(df["senior_id"], errors="coerce")
                        df["value"] = pd.to_numeric(df["value"], errors="coerce")
                        df["sbp"] = pd.to_numeric(df["sbp"], errors="coerce")
                        df["dbp"] = pd.to_numeric(df["dbp"], errors="coerce")
                        df["type"] = df["type"].astype(str).str.strip()
                        df = df.dropna(subset=["senior_id"]).reset_index(drop=True)
                        # Ensure seniors exist to satisfy FK
                        bulk_upsert_seniors(conn, df["senior_id"].dropna().unique())
                        # Append to DB
                        df.to_sql("measurements", conn, if_exists="append", index=False)
                        conn.commit()
                        inserted = len(df)
                        total_inserted += inserted
                        processed += len(rows_buffer)
                        logger.info(f"  {sheet_name}: +{inserted:,} (total {total_inserted:,})")
                        rows_buffer.clear()
                        # Update ingestion_state as in-progress
                        try:
                            cur.execute(
                                "INSERT INTO ingestion_state(source, status, rows_processed) VALUES(?, 'in-progress', ?)\n"
                                "ON CONFLICT(source) DO UPDATE SET status='in-progress', rows_processed=excluded.rows_processed, updated_at=CURRENT_TIMESTAMP",
                                (source_key, total_inserted),
                            )
                            conn.commit()
                        except sqlite3.OperationalError:
                            pass

                # Flush remainder
                if rows_buffer:
                    df = pd.DataFrame.from_records(rows_buffer, columns=["senior_id","value","sbp","dbp","date","type"])
                    df["senior_id"] = pd.to_numeric(df["senior_id"], errors="coerce")
                    df["value"] = pd.to_numeric(df["value"], errors="coerce")
                    df["sbp"] = pd.to_numeric(df["sbp"], errors="coerce")
                    df["dbp"] = pd.to_numeric(df["dbp"], errors="coerce")
                    df["type"] = df["type"].astype(str).str.strip()
                    df = df.dropna(subset=["senior_id"]).reset_index(drop=True)
                    bulk_upsert_seniors(conn, df["senior_id"].dropna().unique())
                    df.to_sql("measurements", conn, if_exists="append", index=False)
                    conn.commit()
                    inserted = len(df)
                    total_inserted += inserted
                    processed += len(rows_buffer)
                    logger.info(f"  {sheet_name}: +{inserted:,} (total {total_inserted:,})")
                    rows_buffer.clear()
                    try:
                        cur.execute(
                            "INSERT INTO ingestion_state(source, status, rows_processed) VALUES(?, 'in-progress', ?)\n"
                            "ON CONFLICT(source) DO UPDATE SET status='in-progress', rows_processed=excluded.rows_processed, updated_at=CURRENT_TIMESTAMP",
                            (source_key, total_inserted),
                        )
                        conn.commit()
                    except sqlite3.OperationalError:
                        pass

                logger.info(f"✓ Finished sheet '{sheet_name}'")
                # Mark as done
                try:
                    cur.execute(
                        "INSERT INTO ingestion_state(source, status, rows_processed) VALUES(?, 'done', ?)\n"
                        "ON CONFLICT(source) DO UPDATE SET status='done', rows_processed=excluded.rows_processed, updated_at=CURRENT_TIMESTAMP",
                        (source_key, total_inserted),
                    )
                    conn.commit()
                except sqlite3.OperationalError:
                    pass
        finally:
            wb.close()
    logger.info(f"✓ Measurements streaming complete. Inserted {total_inserted:,} rows in total.")


def insert_medical_info(df: pd.DataFrame, conn: sqlite3.Connection):
    """Insert medical information."""
    logger.info(f"Inserting {len(df)} medical records...")
    df.to_sql("medical_info", conn, if_exists="append", index=False)
    conn.commit()
    logger.info("✓ Medical info inserted successfully")


def insert_alerts(df: pd.DataFrame, conn: sqlite3.Connection):
    """Insert alert records."""
    logger.info(f"Inserting {len(df)} alert records...")
    df.to_sql("alerts", conn, if_exists="append", index=False)
    conn.commit()
    logger.info("✓ Alerts inserted successfully")


def load_all_data(
    data_dir: Path = RAW_DATA_PATH,
    fresh_start: bool = False,
    streaming: bool = True,
    batch_rows: int = 100_000,
    resume: bool = True,
):
    """
    Main pipeline: Load all data from Excel files into SQLite.
    
    Args:
        data_dir: Directory containing the Excel files
        fresh_start: If True, delete existing database and start fresh
    """
    measurement_files = [
        data_dir / "data_202512221122-01-09.xlsx",
        data_dir / "data_202512221231-16-23.xlsx",
        data_dir / "data_202512221344-24-30.xlsx",
    ]
    medical_file = data_dir / "Med&Diseases_202512221410.xlsx"
    alerts_file = data_dir / "SOS_202512221411.xlsx"
    seniors_demo_file = data_dir / "SeniorGenderAge_202512221409.xlsx"

    db_path = Path(__file__).parent.parent.parent / "db" / "hrp_data.db"
    if fresh_start and db_path.exists():
        db_path.unlink()
        logger.info("Deleted existing database")
    
    conn = initialize_database(db_path)
    
    try:
        # Load seniors demographics first (to satisfy FK during measurements insert)
        if seniors_demo_file.exists():
            seniors_df = load_seniors_demographics(seniors_demo_file)
            insert_seniors_demographics(seniors_df, conn)
        else:
            logger.warning(f"Seniors demographic file not found: {seniors_demo_file}")

        # Load measurements (largest dataset)
        if streaming:
            insert_measurements_from_excel(measurement_files, conn, batch_rows=batch_rows, resume=resume)
        else:
            measurements = load_measurements_data(measurement_files)
            # Ensure seniors exist for FK before insert
            bulk_upsert_seniors(conn, measurements["senior_id"].dropna().unique())
            insert_measurements(measurements, conn)
        
        # Load medical info
        medical_info = load_medical_info(medical_file)
        insert_medical_info(medical_info, conn)
        
        # Load alerts
        alerts = load_alerts(alerts_file)
        insert_alerts(alerts, conn)
        
        # Normalize and sync derived tables from raw imports
        normalize_from_sources(conn)
        
        logger.info("\n✓✓✓ All data loaded and normalized successfully! ✓✓✓")
        
        # Print summary statistics
        print_summary_stats(conn)
        
    finally:
        conn.close()


def print_summary_stats(conn: sqlite3.Connection):
    """Print database summary statistics."""
    cursor = conn.cursor()
    
    def _count(table: str) -> int:
        try:
            return int(cursor.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])
        except sqlite3.OperationalError:
            return 0
    
    stats = {
        "seniors": _count("seniors"),
        "measurements": _count("measurements"),
        "alerts": _count("alerts"),
        "medical_info (raw)": _count("medical_info"),
        "diseases": _count("diseases"),
        "medicines": _count("medicines"),
        "senior_diseases": _count("senior_diseases"),
        "senior_medicines": _count("senior_medicines"),
    }
    
    print("\n" + "="*50)
    print("DATABASE SUMMARY")
    print("="*50)
    for key, value in stats.items():
        print(f"{key:.<30} {value:>15,}")
    print("="*50)


if __name__ == "__main__":
    load_all_data(fresh_start=True)
